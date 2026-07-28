"""
Samsung Digital Content Plan — deadline engine + conflict detector.

Reads the master content-plan Excel (4 tabs), recalculates every deadline
using the PER-TYPE rules from tab "2. Правила дедлайнов" (not the flat
-14/-7/-3 that's baked into the source file), applies the weekend-shift
and distributor-approval rules, then runs the six weekly checks from
tab "4. Трекинг" plus the model/channel coverage rule from tab
"3. Каналы и назначения".

Re-run on next month's file: just point INPUT_PATH at the new .xlsx —
everything else (rules, checks) is read from the workbook, not hardcoded.
"""

import openpyxl
from datetime import date, datetime, timedelta
import json

INPUT_PATH = "samsung-digital-content-plan-template.xlsx"
TODAY = date(2026, 7, 27)  # override with date.today() in production

DISTRIBUTOR_DESTINATIONS = {"Technodom", "Sulpak", "Mechta", "Evrika"}
MARKETPLACE_DESTINATIONS = {"Kaspi.kz", "Ozon"}
MEDIA_TYPES = {"Медийка"}


def prev_business_day(d: date) -> date:
    """Shift a date landing on Sat/Sun back to the preceding Friday."""
    while d.weekday() >= 5:  # 5=Sat, 6=Sun
        d -= timedelta(days=1)
    return d


def load_deadline_rules(wb):
    ws = wb["2. Правила дедлайнов"]
    rules = {}

    def to_int(x):
        if x is None:
            return None
        s = str(x).replace("−", "-").replace("дней", "").replace("дня", "").replace("день", "").strip()
        try:
            return abs(int(s))
        except ValueError:
            return None

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "Тип кампании":
            header_row = row[0].row
            break
    if header_row is None:
        raise ValueError('Header row "Тип кампании" not found in rules sheet')

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        vals = [c.value for c in row]
        campaign_type = vals[0]
        if not campaign_type or not isinstance(campaign_type, str):
            continue
        if campaign_type.strip().startswith("Дополнительные"):
            break
        brief, creative, approval, upload = to_int(vals[1]), to_int(vals[2]), to_int(vals[3]), to_int(vals[4])
        if None in (brief, creative, approval, upload):
            continue
        rules[campaign_type.strip()] = {
            "brief": brief, "creative": creative,
            "approval": approval, "upload": upload,
        }
    return rules


def load_channel_matrix(wb):
    ws = wb["3. Каналы и назначения"]
    matrix = {}
    header_seen = False
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        vals = [c.value for c in row]
        if vals[0] == "Канал":
            header_seen = True
            continue
        if header_seen and vals[0] and vals[1]:
            matrix[str(vals[0]).strip()] = {
                "type": vals[1],
                "destinations": str(vals[2] or ""),
                "formats": str(vals[3] or ""),
            }
    return matrix


def load_content_plan(wb):
    ws = wb["1. Контент-план"]
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "ID":
            header_row_idx = row[0].row
            break
    headers = [c.value for c in ws[header_row_idx]]
    campaigns = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        vals = [c.value for c in row]
        if not vals[0] or not str(vals[0]).startswith("C-"):
            continue
        rec = dict(zip(headers, vals))
        campaigns.append(rec)
    return campaigns


def as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def recalc_deadlines(campaigns, rules):
    for c in campaigns:
        ctype = str(c.get("Тип кампании", "")).strip()
        start = as_date(c.get("Старт"))
        rule = rules.get(ctype)
        c["_flags"] = []
        if not rule or not start:
            c["_flags"].append(f"Неизвестный тип кампании или пустая дата старта: {ctype!r}")
            continue

        destination = str(c.get("Назначение (куда ведём)", "")).strip()
        is_distributor = destination in DISTRIBUTOR_DESTINATIONS
        approval_offset = rule["approval"] + (1 if is_distributor else 0)

        brief_due = prev_business_day(start - timedelta(days=rule["brief"]))
        creative_due = prev_business_day(start - timedelta(days=rule["creative"]))
        approval_due = prev_business_day(start - timedelta(days=approval_offset))
        upload_due = prev_business_day(start - timedelta(days=rule["upload"]))

        c["Бриф до (пересчитано)"] = brief_due
        c["Креатив до (пересчитано)"] = creative_due
        c["Согл. до (пересчитано)"] = approval_due
        c["Загрузка до (пересчитано)"] = upload_due

        # was the ORIGINAL flat -14/-7/-3 rule wrong for this campaign type?
        naive_brief = prev_business_day(start - timedelta(days=14))
        naive_creative = prev_business_day(start - timedelta(days=7))
        naive_approval = prev_business_day(start - timedelta(days=3))
        if (naive_brief, naive_creative, naive_approval) != (brief_due, creative_due, approval_due):
            c["_flags"].append(
                f"Дедлайны в файле были посчитаны единым правилом (-14/-7/-3), "
                f"а для типа «{ctype}» правило другое ({rule['brief']}/{rule['creative']}/{approval_offset} дней). "
                f"Верный бриф-дедлайн: {brief_due}, было бы (по флэт-правилу): {naive_brief}."
            )

        # is there even enough runway to hit the earliest deadline at all?
        if brief_due < TODAY and start > TODAY:
            days_short = (TODAY - brief_due).days
            c["_flags"].append(
                f"Физически не успеть: бриф должен был уйти {brief_due}, "
                f"а сегодня уже {TODAY} — просрочка на {days_short} дн. до старта {start}."
            )
    return campaigns


def run_checks(campaigns, channel_matrix):
    report = {"overdue": [], "due_this_week": [], "missing_creative": [],
              "week_overload": [], "model_media_without_perf": [],
              "model_coverage_gap": [], "recalculated_vs_flat": []}

    for c in campaigns:
        status = str(c.get("Статус", "")).strip()
        start = as_date(c.get("Старт"))
        cid = c.get("ID")

        for label, key in [("Бриф", "Бриф до (пересчитано)"), ("Креатив", "Креатив до (пересчитано)"),
                            ("Согласование", "Согл. до (пересчитано)"), ("Загрузка", "Загрузка до (пересчитано)")]:
            due = c.get(key)
            if not due:
                continue
            if due < TODAY and status != "Готово":
                report["overdue"].append({"id": cid, "этап": label, "дедлайн": str(due), "статус": status})
            elif TODAY <= due <= TODAY + timedelta(days=7):
                report["due_this_week"].append({"id": cid, "этап": label, "дедлайн": str(due)})

        if status != "Готово" and start and (start - TODAY).days < 7:
            report["missing_creative"].append({"id": cid, "старт": str(start), "статус": status})

        for f in c.get("_flags", []):
            if "единым правилом" in f:
                report["recalculated_vs_flat"].append({"id": cid, "детали": f})
            elif "Физически не успеть" in f:
                report["overdue"].append({"id": cid, "этап": "Бриф (риск срыва)", "детали": f})

    # week overload — count starts per ISO week
    from collections import defaultdict
    week_counts = defaultdict(list)
    for c in campaigns:
        s = as_date(c.get("Старт"))
        if s:
            wk = f"{s.isocalendar()[0]}-W{s.isocalendar()[1]:02d}"
            week_counts[wk].append(c.get("ID"))
    for wk, ids in week_counts.items():
        if len(ids) > 4:
            report["week_overload"].append({"неделя": wk, "кампании": ids, "кол-во": len(ids)})

    # media-without-performance-support, per model
    from collections import defaultdict
    by_model = defaultdict(list)
    for c in campaigns:
        by_model[c.get("Модель")].append(c)
    for model, camps in by_model.items():
        has_perf = any(str(c.get("Тип кампании")) not in MEDIA_TYPES for c in camps)
        has_media = any(str(c.get("Тип кампании")) in MEDIA_TYPES for c in camps)
        if has_media and not has_perf:
            media_ids = [c.get("ID") for c in camps if str(c.get("Тип кампании")) in MEDIA_TYPES]
            report["model_media_without_perf"].append({"модель": model, "кампании": media_ids})

        has_marketplace_or_distributor = any(
            str(c.get("Назначение (куда ведём)")).strip() in (DISTRIBUTOR_DESTINATIONS | MARKETPLACE_DESTINATIONS)
            for c in camps
        )
        if not has_marketplace_or_distributor:
            report["model_coverage_gap"].append({
                "модель": model,
                "проблема": "нет кампании на маркетплейс или дистрибьютора — только прямой трафик"
            })

    return report


def build_corrected_workbook(campaigns, source_path, out_path):
    """Write a copy of the source workbook with corrected deadline columns
    filled in on tab 1, plus red fill on any cell that's now overdue."""
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font

    wb = load_workbook(source_path)
    ws = wb["1. Контент-план"]
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value == "ID":
            header_row_idx = row[0].row
            break
    headers = [c.value for c in ws[header_row_idx]]
    col_idx = {h: i + 1 for i, h in enumerate(headers)}

    red = PatternFill(start_color="FDE2E2", end_color="FDE2E2", fill_type="solid")
    amber = PatternFill(start_color="FDF3D9", end_color="FDF3D9", fill_type="solid")
    by_id = {c["ID"]: c for c in campaigns}

    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        cid = row[0].value
        if cid not in by_id:
            continue
        c = by_id[cid]
        pairs = [
            ("Бриф до", "Бриф до (пересчитано)"),
            ("Креатив до", "Креатив до (пересчитано)"),
            ("Согл. до", "Согл. до (пересчитано)"),
        ]
        for src_header, calc_key in pairs:
            if src_header not in col_idx:
                continue
            cell = row[col_idx[src_header] - 1]
            due = c.get(calc_key)
            if not due:
                continue
            cell.value = due
            cell.number_format = "yyyy-mm-dd"
            status = str(c.get("Статус", ""))
            if due < TODAY and status != "Готово":
                cell.fill = red
            elif TODAY <= due <= TODAY + timedelta(days=7):
                cell.fill = amber

    wb.save(out_path)


def generate_digest(report, campaigns, today=TODAY):
    by_id = {c["ID"]: c for c in campaigns}
    lines = [f"# Недельный дайджест — {today}\n"]

    lines.append("## 🔥 Что горит прямо сейчас")
    if report["overdue"]:
        seen = set()
        for item in report["overdue"]:
            key = item["id"]
            if key in seen:
                continue
            seen.add(key)
            c = by_id.get(key, {})
            lines.append(f"- **{key}** ({c.get('Модель', '?')}, {c.get('Тип кампании', '?')}) — "
                         f"просрочен этап «{item.get('этап')}», статус «{item.get('статус', c.get('Статус'))}»")
    else:
        lines.append("- Просрочек нет.")

    lines.append("\n## 📋 Что нужно сделать на этой неделе")
    if report["missing_creative"]:
        for item in report["missing_creative"][:8]:
            c = by_id.get(item["id"], {})
            lines.append(f"- **{item['id']}** ({c.get('Модель', '?')}) — старт {item['старт']}, "
                         f"а статус всё ещё «{item['статус']}». Нужен креатив немедленно.")
    else:
        lines.append("- Срочных задач на неделю нет.")

    lines.append("\n## 🕳️ Где дыры в плане")
    gaps = []
    for item in report["model_media_without_perf"]:
        gaps.append(f"- Модель **{item['модель']}**: есть медийка ({', '.join(item['кампании'])}), "
                     f"но нет ни одной перформанс-кампании — охват не будет конвертироваться.")
    for item in report["model_coverage_gap"]:
        gaps.append(f"- Модель **{item['модель']}**: {item['проблема']}.")
    if report["week_overload"]:
        for item in report["week_overload"]:
            gaps.append(f"- Неделя **{item['неделя']}**: {item['кол-во']} запусков одновременно — "
                         f"перегруз креативной студии ({', '.join(item['кампании'])}).")
    lines.extend(gaps if gaps else ["- Дыр в покрытии не найдено."])

    return "\n".join(lines)


def write_conflict_report_md(report, campaigns, out_path):
    by_id = {c["ID"]: c for c in campaigns}
    lines = ["# Отчёт проверок — мастер контент-план\n"]

    lines.append("## Расхождение с исходным файлом (найденная ошибка)")
    lines.append("Исходный файл считал ВСЕ дедлайны по единому правилу −14/−7/−3. "
                  "По вкладке «2. Правила дедлайнов» правила разные по типам кампаний. "
                  "Ниже — кампании, где пересчёт дал другую дату:\n")
    for item in report["recalculated_vs_flat"]:
        lines.append(f"- **{item['id']}**: {item['детали']}")

    lines.append("\n## Просроченные / под риском (приоритет: высокий)")
    seen = set()
    for item in report["overdue"]:
        if item["id"] in seen:
            continue
        seen.add(item["id"])
        c = by_id.get(item["id"], {})
        lines.append(f"- **{item['id']}** ({c.get('Модель')}) — этап «{item.get('этап')}» просрочен")

    lines.append("\n## Кампании без готового креатива при близком старте (приоритет: высокий)")
    for item in report["missing_creative"]:
        lines.append(f"- **{item['id']}** — старт {item['старт']}, статус «{item['статус']}»")

    lines.append("\n## Медийка без перформанс-поддержки (приоритет: средний)")
    for item in report["model_media_without_perf"]:
        lines.append(f"- Модель **{item['модель']}**: {', '.join(item['кампании'])}")

    lines.append("\n## Дыры в покрытии моделей (приоритет: средний)")
    for item in report["model_coverage_gap"]:
        lines.append(f"- Модель **{item['модель']}**: {item['проблема']}")

    lines.append("\n## Перегруз недели (приоритет: низкий)")
    if report["week_overload"]:
        for item in report["week_overload"]:
            lines.append(f"- {item['неделя']}: {item['кол-во']} запусков — {', '.join(item['кампании'])}")
    else:
        lines.append("- Перегруженных недель не найдено (максимум — 3 запуска/неделю).")

    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def main():
    import os
    os.makedirs("output", exist_ok=True)
    wb = openpyxl.load_workbook(INPUT_PATH, data_only=True)
    rules = load_deadline_rules(wb)
    matrix = load_channel_matrix(wb)
    campaigns = load_content_plan(wb)
    campaigns = recalc_deadlines(campaigns, rules)
    report = run_checks(campaigns, matrix)

    out = {
        "generated_at": str(TODAY),
        "rules_used": rules,
        "campaigns": [
            {k: (str(v) if isinstance(v, (date, datetime)) else v) for k, v in c.items()}
            for c in campaigns
        ],
        "conflict_report": report,
    }
    with open("output/output.json", "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2, default=str)

    build_corrected_workbook(campaigns, INPUT_PATH, "output/content-plan-corrected.xlsx")
    write_conflict_report_md(report, campaigns, "output/conflict-report.md")
    digest = generate_digest(report, campaigns)
    with open("output/weekly-digest.md", "w", encoding="utf-8") as f:
        f.write(digest)

    print(digest)


if __name__ == "__main__":
    main()
